"""Simple Flask web application for attendance system.
Minimal code approach with basic functionality.
"""

import base64
import os
from datetime import datetime
from io import BytesIO

import cv2
import numpy as np
from flask import Flask, jsonify, redirect, render_template, request, url_for, send_file, make_response
from werkzeug.utils import secure_filename

# Handle both relative and absolute imports
try:
    from .attendance_system import AttendanceSystem
    from .config import ALLOWED_EXTENSIONS, WEB_DEBUG, WEB_HOST, WEB_PORT, ATTENDANCE_DIR
except ImportError:
    from attendance_system import AttendanceSystem
    from config import ALLOWED_EXTENSIONS, WEB_DEBUG, WEB_HOST, WEB_PORT, ATTENDANCE_DIR

# Get absolute paths for templates and static folders
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE_DIR = os.path.join(BASE_DIR, "templates")
STATIC_DIR = os.path.join(BASE_DIR, "static")

app = Flask(__name__, template_folder=TEMPLATE_DIR, static_folder=STATIC_DIR)
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16MB max file size

# Initialize attendance system
attendance_system = AttendanceSystem()
# Note: Model selection will happen on first use (lazy loading)


def allowed_file(filename):
    """Check if file extension is allowed."""
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route("/")
def index():
    """Main page with attendance marking options."""
    # Don't initialize models on page load - load asynchronously
    # This makes the page load instantly
    return render_template("index.html", users=[], attendance=[])


@app.route("/add_user")
def add_user():
    """Add new user page."""
    return render_template("add_user.html")


@app.route("/add_user", methods=["POST"])
def add_user_post():
    """Handle new user addition."""
    user_name = request.form.get("user_name")
    uploaded_files = request.files.getlist("user_images")

    if not user_name:
        return jsonify({"success": False, "message": "User name is required"})

    # Save uploaded files temporarily
    temp_files = []
    for file in uploaded_files:
        if file and file.filename and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            temp_path = os.path.join("temp", filename)
            os.makedirs("temp", exist_ok=True)
            file.save(temp_path)
            temp_files.append(temp_path)

    if not temp_files:
        return jsonify({"success": False, "message": "No valid image files provided"})

    # Add user to system
    result = attendance_system.add_new_user(user_name, temp_files)

    # Clean up temporary files
    for temp_file in temp_files:
        try:
            os.remove(temp_file)
        except OSError:
            pass

    if result["success"]:
        return redirect(url_for("index"))
    return jsonify(result)


@app.route("/mark_attendance_camera", methods=["POST"])
def mark_attendance_camera():
    """Mark attendance using camera capture."""
    if request.json is None:
        return jsonify({"success": False, "message": "No JSON data provided"})

    camera_source = request.json.get("camera_source", 0)
    auto_mode = request.json.get("auto_mode", False)
    save_captured = not auto_mode  # Only save in manual mode

    # Convert camera_source to appropriate type
    if isinstance(camera_source, str) and camera_source.isdigit():
        camera_source = int(camera_source)

    # Capture image from camera
    image = attendance_system.capture_from_camera(camera_source)
    if image is None:
        return jsonify({"success": False, "message": "Failed to capture from camera"})

    # Mark attendance
    result = attendance_system.mark_attendance(image, save_captured=save_captured)

    # Convert image to base64 for display
    if result["success"]:
        _, buffer = cv2.imencode(".jpg", image)
        img_base64 = base64.b64encode(buffer).decode("utf-8")
        result["captured_image"] = img_base64

    return jsonify(result)


@app.route("/mark_attendance_upload", methods=["POST"])
def mark_attendance_upload():
    """Mark attendance using uploaded image."""
    if "image" not in request.files:
        return jsonify({"success": False, "message": "No image file provided"})

    file = request.files["image"]
    if file.filename == "" or file.filename is None:
        return jsonify({"success": False, "message": "No image file selected"})

    if not allowed_file(file.filename):
        return jsonify({"success": False, "message": "Invalid file type"})

    # Save uploaded file temporarily
    filename = secure_filename(file.filename)
    temp_path = os.path.join("temp", filename)
    os.makedirs("temp", exist_ok=True)
    file.save(temp_path)

    # Mark attendance
    result = attendance_system.mark_attendance(temp_path)

    # Clean up temporary file
    try:
        os.remove(temp_path)
    except OSError:
        pass

    return jsonify(result)


@app.route("/get_attendance")
def get_attendance():
    """Get attendance records, optionally filtered by date range."""
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    if start_date and end_date:
        # Get attendance for date range
        attendance = get_attendance_by_date_range(start_date, end_date)
    else:
        # Get today's attendance
        attendance = attendance_system.get_today_attendance()
    
    return jsonify(attendance)


@app.route("/get_users")
def get_users():
    """Get list of registered users."""
    users = attendance_system.get_user_list()
    return jsonify(users)


@app.route("/delete_user/<user_name>", methods=["DELETE"])
def delete_user(user_name):
    """Delete a registered user."""
    result = attendance_system.delete_user(user_name)
    return jsonify(result)


@app.route("/camera_test/<path:camera_source>")
def camera_test(camera_source):
    """Test camera functionality."""
    # Convert camera_source to appropriate type
    if camera_source.isdigit():
        camera_source = int(camera_source)

    print(f"🧪 Testing camera: {camera_source}")
    image = attendance_system.capture_from_camera(camera_source)
    if image is not None:
        _, buffer = cv2.imencode(".jpg", image)
        img_base64 = base64.b64encode(buffer).decode("utf-8")
        return jsonify(
            {
                "success": True,
                "image": img_base64,
                "message": "Camera working successfully!",
            },
        )
    error_message = "Camera not available"
    if isinstance(camera_source, str):
        error_message += f" - Check IP camera URL: {camera_source}"
    return jsonify({"success": False, "message": error_message})


# Live video streaming disabled - UI simplified to use only Test Camera and Mark Attendance
# @app.route("/video_feed/<path:camera_source>")
# def video_feed(camera_source):
#     """Video streaming route with face detection and recognition.
#     
#     Args:
#         camera_source: Camera index (int) or IP camera URL (str)
#     
#     Returns:
#         MJPEG stream response with detected faces and labels
#     """
#     # Convert camera_source to appropriate type
#     if camera_source.isdigit():
#         camera_source = int(camera_source)
#     
#     return make_response(
#         generate_video_frames(camera_source),
#         200,
#         {'Content-Type': 'multipart/x-mixed-replace; boundary=frame'}
#     )
# 
# 
# def generate_video_frames(camera_source):
#     """Generate video frames with face detection and bounding boxes."""
#     cap = None
#     try:
#         # For IP cameras (string URLs), use CAP_FFMPEG backend with timeout settings
#         if isinstance(camera_source, str):
#             cap = cv2.VideoCapture(camera_source, cv2.CAP_FFMPEG)
#             
#             # Set timeout properties for network streams (in milliseconds)
#             cap.set(cv2.CAP_PROP_OPEN_TIMEOUT_MSEC, 5000)  # 5 second open timeout
#             cap.set(cv2.CAP_PROP_READ_TIMEOUT_MSEC, 5000)  # 5 second read timeout
#         else:
#             # For local cameras, use default backend
#             cap = cv2.VideoCapture(camera_source)
#         
#         if not cap or not cap.isOpened():
#             print(f"❌ Failed to open camera: {camera_source}")
#             # Return error frame instead of None
#             error_frame = create_error_frame("Camera not available")
#             ret, buffer = cv2.imencode('.jpg', error_frame)
#             if ret:
#                 yield (b'--frame\r\n'
#                        b'Content-Type: image/jpeg\r\n\r\n' + buffer.tobytes() + b'\r\n')
#             return
#         
#         # Set buffer size for better performance
#         cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
#         
#         while True:
#             ret, frame = cap.read()
#             if not ret or frame is None:
#                 print("❌ Failed to read frame")
#                 break
#             
#             # Detect and draw faces with names
#             frame_with_boxes = attendance_system.draw_faces_with_names(frame)
#             
#             # Encode frame as JPEG with balanced quality for streaming
#             ret, buffer = cv2.imencode('.jpg', frame_with_boxes, [cv2.IMWRITE_JPEG_QUALITY, 75])
#             if not ret:
#                 continue
#             
#             # Yield frame in multipart format
#             frame_bytes = buffer.tobytes()
#             yield (b'--frame\r\n'
#                    b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')
#     
#     except GeneratorExit:
#         # Client disconnected
#         print("Client disconnected from video stream")
#     except Exception as e:
#         print(f"❌ Video streaming error: {e}")
#     finally:
#         if cap is not None:
#             cap.release()
#             print("📷 Camera released")
# 
# 
# def create_error_frame(message):
#     """Create an error frame with a message."""
#     # Create a simple error image
#     error_img = np.zeros((480, 640, 3), dtype=np.uint8)
#     error_img[:] = (40, 40, 40)  # Dark gray background
#     
#     # Add error message
#     font = cv2.FONT_HERSHEY_SIMPLEX
#     text_size = cv2.getTextSize(message, font, 1, 2)[0]
#     text_x = (640 - text_size[0]) // 2
#     text_y = (480 + text_size[1]) // 2
#     
#     cv2.putText(error_img, message, (text_x, text_y), font, 1, (0, 0, 255), 2)
#     return error_img


# CNN training page removed - using only custom embedding model for Raspberry Pi optimization
# @app.route("/cnn_training")
# def cnn_training():
#     """CNN training page."""
#     return render_template("cnn_training.html")


@app.route("/model_status")
def model_status():
    """Get current model status (active model and availability)."""
    try:
        status = attendance_system.get_current_model_info()
        return jsonify(status)
    except Exception as e:
        return jsonify({"error": str(e)})


@app.route("/initialize_system", methods=["POST"])
def initialize_system():
    """Initialize the face recognition system in the background."""
    try:
        import sys
        print("=" * 80, file=sys.stderr)
        print("🚀 Starting system initialization...", file=sys.stderr)
        print("=" * 80, file=sys.stderr)
        
        # This will trigger lazy loading of face manager
        print("📋 Step 1/3: Loading user list...", file=sys.stderr)
        users = attendance_system.get_user_list()
        print(f"✅ Step 1/3: Found {len(users)} users", file=sys.stderr)
        
        print("📅 Step 2/3: Loading today's attendance...", file=sys.stderr)
        today_attendance = attendance_system.get_today_attendance()
        print(f"✅ Step 2/3: Found {len(today_attendance)} attendance records", file=sys.stderr)
        
        print("✅ Step 3/3: System initialization complete!", file=sys.stderr)
        print("=" * 80, file=sys.stderr)
        
        return jsonify({
            "success": True,
            "users_count": len(users),
            "attendance_count": len(today_attendance),
            "users": users,
            "attendance": today_attendance
        })
    except Exception as e:
        import traceback
        import sys
        error_msg = str(e)
        error_trace = traceback.format_exc()
        
        print("=" * 80, file=sys.stderr)
        print("❌ INITIALIZATION ERROR:", file=sys.stderr)
        print(error_msg, file=sys.stderr)
        print("-" * 80, file=sys.stderr)
        print(error_trace, file=sys.stderr)
        print("=" * 80, file=sys.stderr)
        
        return jsonify({
            "success": False, 
            "error": f"Failed to initialize system: {error_msg}"
        })



# Model switching disabled - using only custom embedding model for Raspberry Pi
# @app.route("/cnn_switch_model", methods=["POST"])
# def cnn_switch_model():
#     """Switch between CNN, Embedding, and InsightFace models."""
#     try:
#         data = request.get_json()
#         model_type = data.get("model_type", "insightface")
# 
#         if model_type == "cnn":
#             attendance_system.switch_to_cnn_model()
#         elif model_type == "embedding":
#             attendance_system.switch_to_embedding_model()
#         elif model_type in ("custom_embedding", "custom-embedding"):
#             attendance_system.switch_to_custom_embedding_model()
#         else:
#             attendance_system.switch_to_insightface_model()
# 
#         return jsonify({"success": True, "message": f"Switched to {model_type} model"})
#     except Exception as e:
#         return jsonify({"success": False, "message": str(e)})


# Separate switch endpoints removed - using only custom embedding
# @app.route("/switch/insightface", methods=["POST"])
# def switch_insightface():
#     try:
#         attendance_system.switch_to_insightface_model()
#         return jsonify({"success": True, "message": "Switched to insightface model"})
#     except Exception as e:
#         return jsonify({"success": False, "message": str(e)})


# @app.route("/switch/cnn", methods=["POST"])
# def switch_cnn():
#     try:
#         attendance_system.switch_to_cnn_model()
#         return jsonify({"success": True, "message": "Switched to cnn model"})
#     except Exception as e:
#         return jsonify({"success": False, "message": str(e)})


# @app.route("/switch/embedding", methods=["POST"])
# def switch_embedding():
#     try:
#         attendance_system.switch_to_embedding_model()
#         return jsonify({"success": True, "message": "Switched to embedding model"})
#     except Exception as e:
#         return jsonify({"success": False, "message": str(e)})


# @app.route("/switch/custom_embedding", methods=["POST"])
# def switch_custom_embedding():
#     try:
#         attendance_system.switch_to_custom_embedding_model()
#         return jsonify({"success": True, "message": "Switched to custom embedding model"})
#     except Exception as e:
#         return jsonify({"success": False, "message": str(e)})




# CNN training endpoints removed - not needed for Raspberry Pi deployment
# @app.route("/cnn_prepare_data", methods=["POST"])
# def cnn_prepare_data():
#     """Prepare training data."""
#     try:
#         cnn_trainer = attendance_system.get_cnn_trainer()
#         success = cnn_trainer.prepare_training_data()
# 
#         if success:
#             users_count = len(set(cnn_trainer.training_labels))
#             samples_count = len(cnn_trainer.training_data)
#             return jsonify(
#                 {
#                     "success": True,
#                     "message": "Training data prepared successfully",
#                     "users_count": users_count,
#                     "samples_prepared": samples_count,
#                 },
#             )
#         return jsonify({"success": False, "message": "No training data found"})
#     except Exception as e:
#         return jsonify({"success": False, "message": str(e)})
# 
# 
# @app.route("/cnn_train", methods=["POST"])
# def cnn_train():
#     """Train the CNN model."""
#     try:
#         data = request.get_json()
#         epochs = data.get("epochs", 50)
# 
#         cnn_trainer = attendance_system.get_cnn_trainer()
# 
#         # Prepare data first
#         if not cnn_trainer.prepare_training_data():
#             return jsonify({"success": False, "message": "No training data available"})
# 
#         # Train model
#         result = cnn_trainer.train_model(epochs=epochs)
#         return jsonify(result)
# 
#     except Exception as e:
#         return jsonify({"success": False, "message": str(e)})
# 
# 
# @app.route("/cnn_add_training_images", methods=["POST"])
# def cnn_add_training_images():
#     """Add training images for a user."""
#     try:
#         user_name = request.form.get("user_name")
#         uploaded_files = request.files.getlist("images")
# 
#         if not user_name:
#             return jsonify({"success": False, "message": "User name is required"})
# 
#         if not uploaded_files:
#             return jsonify({"success": False, "message": "No images provided"})
# 
#         # Create user directory
#         from config import DATABASE_DIR
# 
#         user_dir = os.path.join(DATABASE_DIR, user_name)
#         os.makedirs(user_dir, exist_ok=True)
# 
#         processed_count = 0
#         for file in uploaded_files:
#             if file and file.filename and allowed_file(file.filename):
#                 filename = secure_filename(file.filename)
#                 # Add timestamp to avoid conflicts
#                 timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
#                 filename = f"{timestamp}_{filename}"
#                 file_path = os.path.join(user_dir, filename)
#                 file.save(file_path)
#                 processed_count += 1
# 
#         if processed_count > 0:
#             # Update face manager database
#             attendance_system.face_manager.add_user_from_database_folder(user_name)
# 
#             return jsonify(
#                 {
#                     "success": True,
#                     "message": f"Added {processed_count} training images",
#                     "images_processed": processed_count,
#                 },
#             )
#         return jsonify({"success": False, "message": "No valid images processed"})
# 
#     except Exception as e:
#         return jsonify({"success": False, "message": str(e)})
# 
# 
# @app.route("/cnn_add_training_video", methods=["POST"])
# def cnn_add_training_video():
#     """Extract training data from video."""
#     try:
#         user_name = request.form.get("user_name")
#         video_file = request.files.get("video")
#         frame_interval = int(request.form.get("frame_interval", 30))
# 
#         if not user_name:
#             return jsonify({"success": False, "message": "User name is required"})
# 
#         if not video_file or not video_file.filename:
#             return jsonify({"success": False, "message": "No video file provided"})
# 
#         # Save video temporarily
#         filename = secure_filename(video_file.filename)
#         temp_path = os.path.join("temp", filename)
#         os.makedirs("temp", exist_ok=True)
#         video_file.save(temp_path)
# 
#         try:
#             # Extract training data
#             cnn_trainer = attendance_system.get_cnn_trainer()
#             result = cnn_trainer.add_training_data_from_video(
#                 temp_path, user_name, frame_interval,
#             )
# 
#             # Update face manager database
#             if result["success"]:
#                 attendance_system.face_manager.add_user_from_database_folder(user_name)
# 
#             return jsonify(result)
# 
#         finally:
#             # Clean up temporary file
#             try:
#                 os.remove(temp_path)
#             except OSError:
#                 pass
# 
#     except Exception as e:
#         return jsonify({"success": False, "message": str(e)})


@app.route("/export_attendance_pdf", methods=["GET"])
def export_attendance_pdf():
    """Export attendance records to PDF."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER
        
        # Get date range from query parameters
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        
        # Get attendance records
        if start_date and end_date:
            attendance_records = get_attendance_by_date_range(start_date, end_date)
            title = f"Attendance Report ({start_date} to {end_date})"
        else:
            attendance_records = attendance_system.get_today_attendance()
            today = datetime.now().strftime("%Y-%m-%d")
            title = f"Attendance Report ({today})"
        
        # Create PDF in memory
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#4F46E5'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor('#6B7280'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        
        # Add title
        elements.append(Paragraph("Face Recognition Attendance System", title_style))
        elements.append(Paragraph(title, subtitle_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Prepare table data
        if attendance_records:
            data = [['#', 'Name', 'Date', 'Time', 'Confidence']]
            for idx, record in enumerate(attendance_records, 1):
                data.append([
                    str(idx),
                    record.get('user_name', 'Unknown'),
                    record.get('date', 'N/A'),
                    record.get('time', 'N/A'),
                    f"{record.get('confidence', 0):.2f}"
                ])
            
            # Create table
            table = Table(data, colWidths=[0.5*inch, 2*inch, 1.5*inch, 1.5*inch, 1.2*inch])
            
            # Table styling
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#111827')),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
                ('TOPPADDING', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ]))
            
            elements.append(table)
            
            # Add summary
            elements.append(Spacer(1, 0.3*inch))
            summary_style = ParagraphStyle(
                'Summary',
                parent=styles['Normal'],
                fontSize=11,
                textColor=colors.HexColor('#374151'),
                spaceAfter=5
            )
            elements.append(Paragraph(f"<b>Total Records:</b> {len(attendance_records)}", summary_style))
            elements.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", summary_style))
        else:
            elements.append(Paragraph("No attendance records found.", styles['Normal']))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        # Send file
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'attendance_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        )
        
    except Exception as e:
        return jsonify({"success": False, "message": f"Error generating PDF: {str(e)}"})


@app.route("/export_attendance_excel", methods=["GET"])
def export_attendance_excel():
    """Export attendance records to Excel."""
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Get date range from query parameters
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        
        # Get attendance records
        if start_date and end_date:
            attendance_records = get_attendance_by_date_range(start_date, end_date)
        else:
            attendance_records = attendance_system.get_today_attendance()
        
        # Create DataFrame
        if attendance_records:
            df = pd.DataFrame(attendance_records)
            # Reorder columns
            columns = ['user_name', 'date', 'time', 'confidence']
            df = df[[col for col in columns if col in df.columns]]
            # Rename columns for better readability
            df.columns = ['Name', 'Date', 'Time', 'Confidence']
        else:
            df = pd.DataFrame(columns=['Name', 'Date', 'Time', 'Confidence'])
        
        # Create Excel file in memory
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Attendance', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Attendance']
            
            # Style header row
            header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Style data cells
            border = Border(
                left=Side(style='thin', color='E5E7EB'),
                right=Side(style='thin', color='E5E7EB'),
                top=Side(style='thin', color='E5E7EB'),
                bottom=Side(style='thin', color='E5E7EB')
            )
            
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Adjust column widths
            column_widths = {'A': 25, 'B': 15, 'C': 15, 'D': 15}
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Add summary information
            summary_row = worksheet.max_row + 2
            worksheet[f'A{summary_row}'] = 'Total Records:'
            worksheet[f'B{summary_row}'] = len(attendance_records)
            worksheet[f'A{summary_row + 1}'] = 'Generated:'
            worksheet[f'B{summary_row + 1}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Make summary bold
            worksheet[f'A{summary_row}'].font = Font(bold=True)
            worksheet[f'A{summary_row + 1}'].font = Font(bold=True)
        
        buffer.seek(0)
        
        # Send file
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'attendance_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        return jsonify({"success": False, "message": f"Error generating Excel: {str(e)}"})


def get_attendance_by_date_range(start_date, end_date):
    """Get attendance records for a date range."""
    import json
    from datetime import datetime, timedelta
    
    records = []
    
    try:
        # Parse dates
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')
        
        # Iterate through date range
        current_date = start
        while current_date <= end:
            date_str = current_date.strftime('%Y-%m-%d')
            attendance_file = os.path.join(ATTENDANCE_DIR, f"attendance_{date_str}.json")
            
            if os.path.exists(attendance_file):
                try:
                    with open(attendance_file) as f:
                        daily_records = json.load(f)
                        records.extend(daily_records)
                except Exception as e:
                    print(f"Error reading {attendance_file}: {e}")
            
            current_date += timedelta(days=1)
        
    except Exception as e:
        print(f"Error getting attendance by date range: {e}")
    
    return records


def run_app():
    """Run the Flask application."""
    print("Starting Simple Attendance System...")
    print(f"Open your browser and go to: http://{WEB_HOST}:{WEB_PORT}")
    print("Press Ctrl+C to stop the server")
    app.run(host=WEB_HOST, port=WEB_PORT, debug=WEB_DEBUG)


if __name__ == "__main__":
    run_app()
